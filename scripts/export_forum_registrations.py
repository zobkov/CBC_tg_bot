#!/usr/bin/env python3
"""
Export bot_forum_registrations (joined with site_registrations) to an xlsx file.

The output matches the column structure of registrations.xlsx plus a user_id column.

Usage:
    python3 scripts/export_forum_registrations.py [output.xlsx]

If no output path is given, the file is saved as:
    storage/forum_registrations_<YYYYMMDD_HHMMSS>.xlsx
"""

import asyncio
import sys
from datetime import datetime
from pathlib import Path

import asyncpg
import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).parent.parent))

from config.config import load_config

# Maps legacy full-name track values (Russian/English) → canonical short key
TRACK_NORMALIZE: dict[str, str] = {
    "Финансы и инвестиции (Finance & Banking)": "finance",
    "Логистика и ВЭД (Supply Chain & Trade)": "logistics",
    "Консалтинг и риск-менеджмент (Strategy & Consulting)": "consulting",
    "Политика, право и дипломатия (Rules of the Game)": "politics",
    "Маркетинг и медиа (Digital & Brand)": "marketing",
    "Язык, культура и перевод (Humanities & Arts)": "language",
    "Китайский трек (Chinese Track)": "chinese",
    "Росмолодежь.Гранты": "rosmolodezh_grants",
    "project_management":"rosmolodezh_grants",
    "Проектный менеджмент (Project Management)":"rosmolodezh_grants"
}

COLUMNS = [
    "user_id",
    "numeric_key",
    "full_name",
    "status",
    "email",
    "adult18",
    "region",
    "participant_status",
    "track",
    "transport",
    "car_number",
    "passport",
    "education",
    "official_invitation",
    "created_at",
]

QUERY = """
    SELECT
        bfr.user_id,
        sr.numeric_key,
        sr.full_name,
        bfr.status,
        sr.email,
        bfr.adult18,
        bfr.region,
        sr.participant_status,
        bfr.track,
        bfr.transport,
        bfr.car_number,
        bfr.passport,
        bfr.education,
        sr.official_invitation,
        bfr.created_at
    FROM bot_forum_registrations bfr
    LEFT JOIN site_registrations sr ON sr.id = bfr.unique_id
    ORDER BY bfr.created_at
"""


async def export(output_path: Path) -> None:
    config = load_config()
    db = config.db

    dsn = (
        f"postgresql://{db.user}:{db.password}"
        f"@{db.host}:{db.port}/{db.database}"
    )

    conn = await asyncpg.connect(dsn=dsn)
    try:
        rows = await conn.fetch(QUERY)
    finally:
        await conn.close()

    if not rows:
        print("No rows returned — xlsx not written.")
        return

    # Group rows by track (preserve insertion order)
    from collections import defaultdict
    by_track: dict[str, list] = defaultdict(list)
    for row in rows:
        track = row["track"] or "no_track"
        track = TRACK_NORMALIZE.get(track, track)
        by_track[track].append(row)

    wb = openpyxl.Workbook()

    def write_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, sheet_rows: list) -> None:
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in sheet_rows:
            values = []
            for col in COLUMNS:
                val = row[col]
                if col == "track" and isinstance(val, str):
                    val = TRACK_NORMALIZE.get(val, val)
                if hasattr(val, "tzinfo") and val.tzinfo is not None:
                    val = val.replace(tzinfo=None)
                values.append(val)
            ws.append(values)

    # First sheet: all rows
    ws_all = wb.active
    ws_all.title = "all"
    write_sheet(ws_all, list(rows))

    # One sheet per track
    for track, track_rows in sorted(by_track.items()):
        # Sheet names have a 31-char limit in xlsx
        sheet_name = track[:31]
        ws_track = wb.create_sheet(title=sheet_name)
        write_sheet(ws_track, track_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    track_summary = ", ".join(f"{t}={len(r)}" for t, r in sorted(by_track.items()))
    print(f"Exported {len(rows)} row(s) → {output_path}")
    print(f"Tracks: {track_summary}")


def main() -> None:
    if len(sys.argv) > 1:
        output_path = Path(sys.argv[1])
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path("storage") / f"forum_registrations_{timestamp}.xlsx"

    asyncio.run(export(output_path))


if __name__ == "__main__":
    main()
